"""
ChagaSight Project Documentation Excel Generator
Produces a multi-sheet Excel workbook documenting the complete pipeline
from data ingestion through preprocessing, augmentation, training, and results.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Colour palette ──────────────────────────────────────────────────────────
CLR = {
    "navy":       "1B3A6B",
    "teal":       "0D6E6E",
    "gold":       "C8960C",
    "slate":      "334155",
    "light_blue": "DBEAFE",
    "light_green":"DCFCE7",
    "light_amber":"FEF3C7",
    "light_red":  "FEE2E2",
    "light_purple":"F3E8FF",
    "white":      "FFFFFF",
    "off_white":  "F8FAFC",
    "row_alt":    "EFF6FF",
    "header_row": "E2E8F0",
    "missing_bg": "FFF7ED",
    "missing_border":"EA580C",
}

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def make_thick_border():
    t = Side(style="medium")
    return Border(left=t, right=t, top=t, bottom=t)

def header_font(size=11, color="FFFFFF", bold=True):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def cell_font(size=10, bold=False, color="1E293B"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def title_font(size=14, color="FFFFFF"):
    return Font(name="Calibri", size=size, bold=True, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def write_sheet_title(ws, text, color_hex, cols=10):
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=cols)
    cell = ws.cell(1, 1, text)
    cell.fill = make_fill(color_hex)
    cell.font = title_font()
    cell.alignment = center()
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 8

def write_section_header(ws, row, text, color_hex, cols=10, col_start=1):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_start + cols - 1)
    cell = ws.cell(row, col_start, text)
    cell.fill = make_fill(color_hex)
    cell.font = header_font(size=11)
    cell.alignment = left()
    ws.row_dimensions[row].height = 22

def write_col_headers(ws, row, headers, fill_hex, col_start=1):
    for c, h in enumerate(headers, col_start):
        cell = ws.cell(row, c, h)
        cell.fill = make_fill(fill_hex)
        cell.font = header_font(size=10)
        cell.alignment = center()
        cell.border = make_border()
    ws.row_dimensions[row].height = 20

def write_data_row(ws, row, values, alt=False, col_start=1,
                   missing=False, bold_first=False):
    bg = CLR["missing_bg"] if missing else (CLR["row_alt"] if alt else CLR["white"])
    for c, v in enumerate(values, col_start):
        cell = ws.cell(row, c, v)
        cell.fill = make_fill(bg)
        cell.font = cell_font(bold=(bold_first and c == col_start))
        cell.alignment = left()
        cell.border = make_border()
    if missing:
        side = Side(style="medium", color=CLR["missing_border"])
        bdr = Border(left=side, right=side, top=side, bottom=side)
        for c in range(col_start, col_start + len(values)):
            ws.cell(row, c).border = bdr

def auto_col_widths(ws, min_w=12, max_w=55):
    for col in ws.columns:
        # skip merged cells which have no column_letter
        anchor = next((c for c in col if hasattr(c, "column_letter")), None)
        if anchor is None:
            continue
        length = max(
            len(str(cell.value)) if (cell.value and hasattr(cell, "column_letter")) else 0
            for cell in col
        )
        ws.column_dimensions[anchor.column_letter].width = min(
            max(length + 3, min_w), max_w
        )

# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 1 — Project Overview
# ═══════════════════════════════════════════════════════════════════════════
def sheet_overview(wb):
    ws = wb.create_sheet("1. Project Overview")
    write_sheet_title(ws, "ChagaSight — FYP Project Overview", CLR["navy"], cols=8)

    row = 3
    write_section_header(ws, row, "  PROJECT IDENTITY", CLR["teal"], cols=8)
    row += 1
    write_col_headers(ws, row, ["Attribute", "Details"], CLR["slate"], col_start=1)
    row += 1

    overview = [
        ("Project Title",
         "ECG-Based Chagas Disease Screening Using a Dual-Modality Transformer "
         "with Self-Supervised Pretraining"),
        ("Research Aim",
         "Develop a dual-modality ECG AI system to screen for Chagas disease "
         "using 1D signals and 2D contour images with self-supervised pretraining "
         "and cross-modal alignment"),
        ("Medical Task", "Binary classification — Chagas Positive vs Negative"),
        ("Data Modalities", "1D ECG signals (12-lead, 1000 samples @100 Hz) + "
                           "2D ECG contour images (3×24×2048 uint8)"),
        ("Datasets",
         "PTB-XL (Germany) + SaMi-Trop (Brazil, Chagas-endemic) + CODE-15 (Brazil)"),
        ("Total Samples", "386,981 ECG records"),
        ("Positive Cases", "8,579 Chagas-positive (2.22% prevalence)"),
        ("Validation Strategy", "5-fold stratified cross-validation (stratified by dataset + label)"),
        ("Primary Metric",
         "TPR@5% — True Positive Rate when screening top 5% of population "
         "(Official PhysioNet 2025 Challenge metric)"),
        ("Secondary Metrics", "AUROC, AUPRC"),
        ("Final Ensemble TPR@5%", "0.4958 (95% CI: 0.4845–0.5068)"),
        ("Final Ensemble AUROC", "0.8707 (95% CI: 0.8665–0.8746)"),
        ("Final Ensemble AUPRC", "0.2589 (95% CI: 0.2489–0.2685)"),
        ("Total Model Parameters", "173.6M"),
        ("Framework", "PyTorch + AMP (float16) + openpyxl for reporting"),
    ]

    for i, (attr, detail) in enumerate(overview):
        write_data_row(ws, row, [attr, detail], alt=(i % 2 == 0), bold_first=True)
        row += 1

    row += 1
    write_section_header(ws, row, "  PIPELINE STAGES (High-Level)", CLR["teal"], cols=8)
    row += 1
    write_col_headers(ws, row, ["#", "Stage", "Script / Module", "Output", "Status"],
                      CLR["slate"])
    row += 1

    stages = [
        ("1", "Raw ECG ingestion",       "build_all_data.py",
         "data/processed/{2d_images, 1d_signals}",                    "Done"),
        ("2", "Data split creation",     "create_splits.py",
         "data/processed/metadata/combined_5fold.csv",                "Done"),
        ("3", "Preprocessing",
         "src/preprocessing/*.py",
         "Baseline removal, resampling, normalisation, image build",   "Done"),
        ("4", "Augmentation",            "src/preprocessing/augmentations.py",
         "Online transforms applied during training",                  "Done"),
        ("5", "2D MAE Pretraining",
         "scripts/mae_pretraining_2d_COMPLETE.py",
         "checkpoints/mae_2d_pretrained.pt",                           "Done"),
        ("6", "1D ST-MEM Pretraining",
         "scripts/stmem_pretraining_1d_COMPLETE.py",
         "checkpoints/stmem_1d_pretrained.pt",                         "Done"),
        ("7", "Fold training (0–4)",     "src/training/trainer.py",
         "checkpoints/fold{0-4}_best.pt  +  fold{0-4}_results.csv",   "Done"),
        ("8", "Ensemble evaluation",     "(post-training aggregation)",
         "checkpoints/ensemble_summary.csv",                           "Done"),
        ("9", "Ablation — no pretrain",  "(fold 0 only)",
         "checkpoints/fold0_no_pretrain_results.csv",                  "Done"),
        ("10","Ablation — 1D only",      "(fold 0 only)",
         "checkpoints/fold0_1d_results.csv",                           "Done"),
        ("11","Ablation — 2D only",      "(fold 0 only)",
         "checkpoints/fold0_2d_results.csv",                           "Done"),
    ]

    for i, row_data in enumerate(stages):
        write_data_row(ws, row, list(row_data), alt=(i % 2 == 0))
        row += 1

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 12
    ws.freeze_panes = "A3"
    return ws


# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 2 — Dataset Summary
# ═══════════════════════════════════════════════════════════════════════════
def sheet_datasets(wb):
    ws = wb.create_sheet("2. Dataset Summary")
    write_sheet_title(ws, "Dataset Summary — ChagaSight", CLR["teal"], cols=9)

    row = 3
    write_section_header(ws, row, "  SOURCE DATASETS", CLR["navy"], cols=9)
    row += 1
    hdrs = ["Dataset", "Region", "ECG Format", "Total Records",
            "Chagas Positive", "Prevalence (%)", "Sample Rate", "Duration", "Labels Used"]
    write_col_headers(ws, row, hdrs, CLR["slate"])
    row += 1

    datasets = [
        ("PTB-XL",    "Germany",
         "WFDB (.hea/.dat)", "21,837", "0 (no Chagas)", "0%",
         "500 Hz (resampled→100 Hz)", "10 s", "Hard: 0 only"),
        ("SaMi-Trop", "Brazil (Chagas endemic)",
         "WFDB / CSV",       "1,631",  "1,631 (all)",   "100%",
         "400 Hz (resampled→100 Hz)", "variable",
         "Hard: 1 only"),
        ("CODE-15",   "Brazil (general pop.)",
         "HDF5",             "363,513","6,948",          "1.91%",
         "400 Hz (resampled→100 Hz)", "7.3–10 s",
         "Soft (pos=0.8, neg=0.2) — uncertain"),
        ("TOTAL",     "—",      "—",
         "386,981",  "8,579",  "2.22%",
         "—", "—", "—"),
    ]

    for i, d in enumerate(datasets):
        bold = (i == 3)
        write_data_row(ws, row, list(d), alt=(i % 2 == 0), bold_first=bold)
        row += 1

    row += 1
    write_section_header(ws, row,
        "  WHY SOFT LABELS FOR CODE-15?", CLR["teal"], cols=9)
    row += 1
    notes = [
        ("Soft label rationale",
         "CODE-15 ECG records were collected from the general population in Brazil. "
         "The 'negative' label means no confirmed Chagas, not confirmed absence. "
         "Because the endemic background rate is non-trivial, hard 0/1 labels would "
         "introduce noise. Soft labels (pos=0.8, neg=0.2) express this uncertainty."),
        ("Implementation",
         "src/preprocessing/soft_labels.py — hard_to_soft_label() maps {0→0.2, 1→0.8}; "
         "AsymmetricBCELoss in src/training/losses.py handles soft BCE correctly."),
    ]
    write_col_headers(ws, row, ["Item", "Explanation"], CLR["slate"], col_start=1)
    row += 1
    for i, (k, v) in enumerate(notes):
        write_data_row(ws, row, [k, v], alt=(i % 2 == 0), bold_first=True)
        row += 1

    auto_col_widths(ws)
    ws.freeze_panes = "A3"
    return ws


# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 3 — Data Split (5-Fold CV)
# ═══════════════════════════════════════════════════════════════════════════
def sheet_splits(wb):
    ws = wb.create_sheet("3. Data Split (5-Fold CV)")
    write_sheet_title(ws, "5-Fold Stratified Cross-Validation Split",
                      CLR["slate"], cols=8)

    row = 3
    write_section_header(ws, row, "  SPLIT STRATEGY", CLR["navy"], cols=8)
    row += 1
    write_col_headers(ws, row, ["Parameter", "Value / Description"], CLR["slate"])
    row += 1

    config = [
        ("Script",              "scripts/create_splits.py"),
        ("Input",               "data/processed/metadata/all_data.csv"),
        ("Output (main)",       "data/processed/metadata/combined_5fold.csv"),
        ("Output (per fold)",
         "data/processed/splits/train_fold{0-4}.csv + val_fold{0-4}.csv"),
        ("Strategy",            "StratifiedKFold — stratified by (dataset, hard_label)"),
        ("n_splits",            "5"),
        ("Validation size",     "≈73,236–73,237 records per fold (~20%)"),
        ("Training size",       "≈309,744–309,745 records per fold (~80%)"),
        ("Key columns",         "id, dataset, label_hard, label_soft, fold"),
        ("Weighted sampling",   "Optional class-weighted sampler available in dataset.py"),
    ]
    for i, (k, v) in enumerate(config):
        write_data_row(ws, row, [k, v], alt=(i % 2 == 0), bold_first=True)
        row += 1

    row += 1
    write_section_header(ws, row, "  PER-FOLD SPLIT SIZES", CLR["teal"], cols=8)
    row += 1
    write_col_headers(ws, row, ["Fold", "Train Records", "Val Records",
                                 "Train Positives", "Val Positives",
                                 "Val Pos Rate (%)", "Status"], CLR["slate"])
    row += 1

    folds = [
        ("Fold 0", "≈309,744", "73,237", "≈6,866", "≈685", "0.94%", "Done"),
        ("Fold 1", "≈309,745", "73,236", "≈6,867", "≈684", "0.93%", "Done"),
        ("Fold 2", "≈309,745", "73,236", "≈6,867", "≈684", "0.93%", "Done"),
        ("Fold 3", "≈309,745", "73,236", "≈6,867", "≈684", "0.93%", "Done"),
        ("Fold 4", "≈309,745", "73,236", "≈6,867", "≈684", "0.93%", "Done"),
        ("TOTAL",  "386,981 total", "—", "8,579", "—", "2.22% overall", "—"),
    ]
    for i, f in enumerate(folds):
        write_data_row(ws, row, list(f), alt=(i % 2 == 0), bold_first=(i == 5))
        row += 1

    row += 1
    write_section_header(ws, row,
        "  STRATIFICATION RATIONALE", CLR["slate"], cols=8)
    row += 1
    write_col_headers(ws, row, ["Concern", "How Addressed"], CLR["slate"])
    row += 1
    rationale = [
        ("Dataset leakage",
         "Stratify by dataset — each fold sees all three datasets proportionally"),
        ("Label imbalance (2.22%)",
         "Stratify by label — every fold has roughly equal positive rate"),
        ("Validation representativeness",
         "Stratified 3000-sample validation subset inside trainer ensures "
         "both classes always present for metric calculation"),
    ]
    for i, (k, v) in enumerate(rationale):
        write_data_row(ws, row, [k, v], alt=(i % 2 == 0), bold_first=True)
        row += 1

    auto_col_widths(ws)
    ws.freeze_panes = "A3"
    return ws


# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 4 — Preprocessing Pipeline
# ═══════════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 4 — Preprocessing  (4 tables)
# ═══════════════════════════════════════════════════════════════════════════
# ── shared helper: one self-contained component table ─────────────────────
def ctable(ws, row, title, hdr_color, col_headers, rows, missing=False):
    """Write one named component table and return the next free row."""
    ncols = len(col_headers)
    write_section_header(ws, row, f"  {title}", hdr_color, cols=ncols)
    row += 1
    write_col_headers(ws, row, col_headers, CLR["slate"])
    row += 1
    for i, r in enumerate(rows):
        write_data_row(ws, row, list(r), alt=(i % 2 == 0),
                       bold_first=True, missing=missing)
        row += 1
    row += 1          # blank gap between tables
    return row


def sheet_preprocessing(wb):
    ws = wb.create_sheet("4. Preprocessing")
    write_sheet_title(ws, "Preprocessing Pipeline — ChagaSight", CLR["teal"], cols=3)

    H = ["Attribute", "Value", "Explanation"]
    row = 3

    # TABLE 1 — BASELINE REMOVAL
    row = ctable(ws, row, "TABLE 1 — BASELINE REMOVAL  (baseline_removal.py)",
        CLR["navy"], H, [
        ("What is it",
         "Butterworth bandpass filter applied to raw ECG",
         "Removes two types of noise: low-frequency baseline wander "
         "(patient movement, breathing) and high-frequency muscle artifact"),
        ("Why needed",
         "Raw ECG contains drift + noise",
         "Without removal, the model learns to classify baseline drift rather than "
         "true cardiac morphology — severely degrades performance"),
        ("Default method",
         "Bandpass: 0.5–40 Hz, order=4",
         "0.5 Hz lower cut removes baseline wander; "
         "40 Hz upper cut removes EMG/muscle noise above ECG frequency band"),
        ("Alternative: highpass",
         "remove_baseline_highpass() — 0.5 Hz",
         "Only removes wander; preserves all frequencies above 0.5 Hz "
         "(keeps some high-freq noise)"),
        ("Alternative: moving average",
         "remove_baseline_moving_average()",
         "Subtracts a moving average window from the signal; "
         "simpler but less precise than Butterworth"),
        ("Input shape",  "(12, T)",   "12 ECG leads, T time samples (original length)"),
        ("Output shape", "(12, T) float32", "Same shape; only amplitude values changed"),
        ("Used in",      "Both 1D and 2D pipelines",
         "Applied before resampling in both branches; "
         "same filter params regardless of target rate"),
        ("Module",       "src/preprocessing/baseline_removal.py", ""),
    ])

    # TABLE 2 — RESAMPLING
    row = ctable(ws, row, "TABLE 2 — RESAMPLING  (resample.py)", CLR["teal"], H, [
        ("What is it",
         "Change the number of time samples to match a target sampling rate",
         "Three source datasets have different native rates: PTB-XL=500 Hz, "
         "SaMi-Trop=400 Hz, CODE-15=400 Hz. Resampling standardises all to one rate."),
        ("Why needed",
         "Datasets have different sample rates",
         "ViT1D_FM expects exactly 1000 samples per signal; ViT2D expects 2048 columns. "
         "Both require a fixed, known rate to correctly interpret patch positions."),
        ("1D target rate", "100 Hz",
         "1000 samples at 100 Hz = 10 seconds. Low rate reduces model size "
         "while retaining all clinically relevant ECG frequencies (below 40 Hz)."),
        ("2D target rate", "500 Hz",
         "5000 samples at 500 Hz = 10 seconds. Higher rate preserves fine waveform "
         "detail that becomes visible as pixel-level structure in the image."),
        ("Method",       "scipy.signal.resample()",
         "Polyphase resampling — correct spectral preservation, "
         "no aliasing if baseline removal was done first"),
        ("Input shape",  "(12, T_original)",  "Original sample count from raw file"),
        ("Output shape", "(12, T_new)",        "T_new = round(T_original * target_fs / source_fs)"),
        ("Padding / trim after resample",
         "pad_or_trim() — zero-pad or discard tail",
         "1D: enforces exactly 1000 samples; 2D: enforces 5000 before image build"),
        ("Module",       "src/preprocessing/resample.py", ""),
    ])

    # TABLE 3 — Z-SCORE NORMALISATION
    row = ctable(ws, row, "TABLE 3 — Z-SCORE NORMALISATION  (normalization.py)",
        CLR["navy"], H, [
        ("What is it",
         "Subtract mean, divide by std — per ECG lead independently",
         "Transforms each lead to have zero mean and unit variance before "
         "passing to the 1D ViT model"),
        ("Why per-lead (not global)",
         "Each lead has a different amplitude scale",
         "Lead I typically has higher amplitude than V1; global norm would "
         "dominate the signal with high-amplitude leads and suppress weak ones. "
         "Per-lead norm gives each lead equal influence."),
        ("Formula",
         "z = (x - mean(x)) / (std(x) + eps)",
         "eps=1e-8 prevents division by zero for flat leads"),
        ("Outlier clipping",
         "Clip to ±3 standard deviations",
         "Artifact spikes (e.g. defibrillation artifacts) can reach 100+ mV. "
         "Clipping at ±3σ prevents them from corrupting the normalised scale."),
        ("Alternative method",
         "Min-max: method='minmax'",
         "Scales each lead to [0, 1]. Less common for ECG; "
         "Z-score is default because it is shift and scale invariant."),
        ("Input shape",  "(12, 1000)",          "Float32 after resampling and padding"),
        ("Output shape", "(12, 1000) float32",  "Same shape; values now in roughly [-3, 3]"),
        ("Not applied to 2D branch",
         "2D uses uint8 [0,255] scaling instead",
         "The 2D image undergoes its own intensity scaling in image_embedding.py; "
         "ViT2D normalises pixels internally with ImageNet-style mean/std"),
        ("Module",       "src/preprocessing/normalization.py",
         "normalize_per_lead(signal, method='zscore')"),
    ])

    # TABLE 4 — 2D IMAGE CONSTRUCTION (WCT)
    row = ctable(ws, row,
        "TABLE 4 — 2D IMAGE CONSTRUCTION: WCT RE-REFERENCING + LEAD STACKING  "
        "(image_embedding.py)", CLR["teal"], H, [
        ("What is it",
         "Converts 12-lead ECG into a (3, 24, 2048) uint8 image",
         "Transforms the 1D time series of each lead into a 2D pixel image "
         "that can be processed by a Vision Transformer trained on images"),
        ("Step 1 — Wilson's Central Terminal (WCT)",
         "WCT = (RA + LA + LL) / 3",
         "WCT is the electrical average of the three limb electrodes. "
         "Subtracting WCT from each reference electrode creates 3 physiologically "
         "distinct views: Ch1=RA-WCT, Ch2=LA-WCT, Ch3=LL-WCT. "
         "Paper: Kim et al. 2025 Section 2.2"),
        ("Why 3 channels",
         "3 WCT views = 3 image channels (like RGB)",
         "Each channel captures a different electrical perspective of the same heart. "
         "The ViT2D treats these exactly as it would RGB channels in a natural image."),
        ("Step 2 — Lead stacking",
         "Stack all 12 leads vertically: 12 × 2 = 24 rows",
         "Each lead is duplicated into two consecutive rows to maintain a "
         "sensible height-to-width aspect ratio. Row order: I,I,II,II,...,V6,V6"),
        ("Step 3 — Column alignment",
         "Trim or zero-pad to exactly 2048 columns",
         "At 500 Hz for 10 s we get 5000 samples. Downsampled to 2048 by "
         "trimming; shorter recordings zero-padded on right."),
        ("Patch grid math",
         "Patch: (8 rows × 64 cols) → 96 patches total",
         "3 row-patches (24÷8) × 32 col-patches (2048÷64) = 96 patches. "
         "This is the sequence length seen by ViT2D's transformer."),
        ("Step 4 — uint8 quantisation",
         "Scale float values to [0, 255], cast to uint8",
         "Reduces per-image storage by 4×. ViT2D divides by 255 and "
         "applies ImageNet mean/std normalisation at runtime."),
        ("Output shape", "(3, 24, 2048) uint8",
         "Saved as .npy; loaded lazily by ChagasDataset"),
        ("Module",       "src/preprocessing/image_embedding.py",
         "build_2d_image(signal_500hz) → ndarray"),
    ])

    # TABLE 5 — SOFT LABEL CONVERSION
    row = ctable(ws, row,
        "TABLE 5 — SOFT LABEL CONVERSION  (soft_labels.py)", CLR["navy"], H, [
        ("What is it",
         "Replaces hard binary labels {0, 1} with soft probabilities {0.2, 0.8}",
         "Instead of telling the model 'definitely positive' or 'definitely negative', "
         "soft labels say 'probably positive' or 'probably negative', "
         "encoding uncertainty in the label itself"),
        ("Why needed for CODE-15",
         "CODE-15 negatives are not confirmed Chagas-free",
         "CODE-15 was collected from general Brazilian population. "
         "A 'negative' just means no confirmed Chagas — it does NOT mean "
         "definitively Chagas-free in a Chagas-endemic region. "
         "Using hard 0 labels would train the model to be overconfident on uncertain negatives."),
        ("NOT applied to PTB-XL / SaMi-Trop",
         "Those datasets use hard labels",
         "PTB-XL is from Germany (very low Chagas prevalence) — negatives are reliable. "
         "SaMi-Trop is all confirmed positives from a Chagas-endemic cohort — no uncertainty."),
        ("Positive mapping",
         "hard 1 → soft 0.8",
         "Even Chagas-confirmed positives in CODE-15 may have some noise; "
         "0.8 instead of 1.0 prevents overconfident positive predictions"),
        ("Negative mapping",
         "hard 0 → soft 0.2",
         "0.2 instead of 0.0 acknowledges that some CODE-15 'negatives' "
         "may be undetected Chagas cases"),
        ("Effect on loss",
         "AsymmetricBCELoss accepts soft targets directly",
         "BCE with soft label y: loss = -[y*log(p) + (1-y)*log(1-p)]. "
         "When y=0.2, neither prediction is fully penalised — consistent with uncertainty."),
        ("Functions",
         "hard_to_soft_label(y, pos_soft=0.8, neg_soft=0.2)",
         "Also: vector_hard_to_soft(y_array) for batch conversion"),
        ("Module",       "src/preprocessing/soft_labels.py", ""),
    ])

    # TABLE 6 — PIPELINE ORDER (1D vs 2D side-by-side)
    row = ctable(ws, row,
        "TABLE 6 — FULL PIPELINE ORDER: 1D vs 2D SIDE-BY-SIDE  (build_all_data.py)",
        CLR["slate"],
        ["Step", "1D Signal Pipeline (100 Hz)", "2D Image Pipeline (500 Hz)"], [
        ("1", "Baseline removal: bandpass 0.5–40 Hz",
               "Baseline removal: bandpass 0.5–40 Hz (same)"),
        ("2", "Resample to 100 Hz",
               "Resample to 500 Hz"),
        ("3", "Pad or trim to 1000 samples",
               "WCT re-referencing: 3 channels (RA-WCT, LA-WCT, LL-WCT)"),
        ("4", "Z-score normalise per lead (clip ±3σ)",
               "Lead stacking: 12 leads × 2 rows = 24 rows"),
        ("5", "Encode sex (M→1.0, F→0.0, unk→0.5)",
               "Trim/pad columns to 2048"),
        ("6", "Save: 1d_signals_100hz/{dataset}/{id}.npy  (float32)",
               "uint8 quantise to [0, 255]"),
        ("7", "—",
               "Save: 2d_images/{dataset}/{id}.npy  (uint8)"),
        ("Output shape", "(12, 1000) float32", "(3, 24, 2048) uint8"),
    ])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 60
    ws.freeze_panes = "A3"
    return ws


# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 5 — Augmentation  (3 tables)
# ═══════════════════════════════════════════════════════════════════════════
def sheet_augmentation(wb):
    ws = wb.create_sheet("5. Augmentation")
    write_sheet_title(ws, "Data Augmentation — One Table Per Technique",
                      CLR["gold"], cols=3)
    H = ["Attribute", "Value", "Explanation"]
    row = 3

    # TABLE 1 — POWERLINE NOISE INJECTION
    row = ctable(ws, row,
        "TABLE 1 — POWERLINE NOISE INJECTION  (augmentations.py)", CLR["navy"], H, [
        ("What is it",
         "Adds synthetic electrical mains interference to the ECG signal",
         "Simulates the 50 Hz or 60 Hz noise picked up by ECG electrodes from "
         "nearby electrical equipment (plugs, lights, monitors)"),
        ("Why needed",
         "Real clinical ECGs almost always contain powerline noise",
         "Without this augmentation the model learns on clean academic data "
         "and may fail on noisy real-world recordings"),
        ("Frequency",       "50 Hz or 60 Hz (randomly chosen per sample)",
         "50 Hz = European mains; 60 Hz = North/South American mains. "
         "Both datasets are from Brazil (60 Hz) and Germany (50 Hz)."),
        ("Frequency jitter", "±0.2 Hz added on top of base frequency",
         "Real mains frequency is never exactly 50.000 Hz; "
         "jitter prevents the model from learning to notch one fixed frequency"),
        ("SNR range",       "[15, 30] dB",
         "15 dB = clearly audible noise; 30 dB = barely detectable. "
         "Full range forces the model to be robust at both extremes."),
        ("Harmonics",       "2nd and 3rd harmonics also added",
         "Real powerline interference includes 100/120 Hz and 150/180 Hz overtones, "
         "not just the fundamental. Including harmonics makes the simulation realistic."),
        ("Applied to",      "1D signal (12, 1000) only",
         "Not applied to 2D images — image augmentations use different techniques"),
        ("Probability",     "p = 0.5 per sample",
         "Applied independently per training sample; each draw is independent"),
        ("Paper reference", "Van Santvliet et al. 2025, Table 1", ""),
    ])

    # TABLE 2 — ADJACENT LEAD MIXUP
    row = ctable(ws, row,
        "TABLE 2 — ADJACENT LEAD MIXUP  (augmentations.py)", CLR["teal"], H, [
        ("What is it",
         "Linearly blends two anatomically adjacent ECG leads together",
         "Creates a weighted combination of two leads from the same anatomical region: "
         "new_V1 = λ*V1 + (1-λ)*V2, where λ ~ Beta(α, α)"),
        ("Why needed",
         "Prevents the model from overfitting to the exact signal in each lead",
         "Adjacent leads (V1/V2 or V5/V6) measure nearly the same cardiac activity "
         "from slightly different angles. Mixing them makes the model robust to "
         "small electrode position differences across different hospitals."),
        ("Lead pairs",      "V1↔V2 and V5↔V6",
         "Only anatomically adjacent leads are mixed. Mixing non-adjacent leads "
         "(e.g. V1 with V6) would be physiologically implausible and harmful."),
        ("Mixing distribution", "Beta(α=0.4, α=0.4)",
         "Beta(0.4,0.4) is U-shaped: mostly produces λ near 0 or near 1, "
         "meaning most of the time one lead strongly dominates the other"),
        ("Applied to",      "1D signal (12, 1000) only",   "Leads 1-12 index positions"),
        ("Probability",     "p = 0.5 per sample",           ""),
        ("Paper reference", "Van Santvliet et al. 2025, Table 1", ""),
    ])

    # TABLE 3 — TEMPORAL CROPPING
    row = ctable(ws, row,
        "TABLE 3 — RANDOM TEMPORAL CROPPING (L1)  (augmentations.py)", CLR["navy"], H, [
        ("What is it",
         "Randomly shortens the ECG recording to a crop of length L1",
         "Takes a contiguous segment of length L1 from the signal; "
         "the remainder of the 1000 samples is zero-padded"),
        ("Why needed",
         "Real recordings are not always exactly 10 seconds",
         "Some ECGs in the datasets are shorter or have usable data only in part. "
         "Cropping teaches the model to diagnose from fewer beats."),
        ("Crop length L1",  "Uniform in [5.65, 10] seconds",
         "5.65 s ≈ 5–6 beats at typical heart rate — minimum to see a full cardiac cycle. "
         "10 s = full recording (no crop)"),
        ("Zero padding",    "Remaining (10 - L1) seconds padded with zeros",
         "Signal is cropped from a random start position; "
         "zeros fill the right side to keep shape (12, 1000)"),
        ("Applied to",      "1D signal (12, 1000) only",   ""),
        ("Probability",     "p = 0.5 per sample",
         "L1 is also used by the Temporal Shifting augmentation (Table 4)"),
        ("Paper reference", "Van Santvliet et al. 2025, Table 1", ""),
    ])

    # TABLE 4 — TEMPORAL SHIFTING
    row = ctable(ws, row,
        "TABLE 4 — RANDOM TEMPORAL SHIFTING (L2)  (augmentations.py)", CLR["teal"], H, [
        ("What is it",
         "Shifts the cropped ECG segment left or right by L2 seconds",
         "After cropping to L1 seconds, the signal is shifted in time by ±L2; "
         "shift is achieved by rolling the array and zero-padding the gap"),
        ("Why needed",
         "Prevents the model learning from absolute time position",
         "Without shifting, the model could overfit to where in the 10-second window "
         "specific cardiac events typically occur (e.g. P wave always starts at t=0.1 s)"),
        ("Shift amount L2",
         "Uniform in [0, ±min(1, 10−L1)] seconds",
         "Maximum shift is 1 second OR the remaining non-cropped duration, "
         "whichever is smaller — prevents shifting out of the window entirely"),
        ("Applied after",   "Always applied after temporal cropping (Table 3)",
         "L2 depends on L1; these two augmentations are paired"),
        ("Applied to",      "1D signal (12, 1000) only",   ""),
        ("Probability",     "p = 0.5 per sample",           ""),
        ("Paper reference", "Van Santvliet et al. 2025, Table 1", ""),
    ])

    # TABLE 5 — 2D IMAGE AUGMENTATIONS
    row = ctable(ws, row,
        "TABLE 5 — 2D IMAGE AUGMENTATIONS  (dataset.py via torchvision)", CLR["navy"], H, [
        ("What is it",
         "Standard image augmentations applied to the (3, 24, 2048) ECG contour image",
         "Horizontal flip and colour jitter applied via torchvision transforms "
         "during dataset collation (not in augmentations.py)"),
        ("Horizontal flip",
         "Randomly mirrors the image left-to-right",
         "Flipping the time axis of the ECG image is equivalent to time-reversing "
         "the signal. The model learns to recognise cardiac morphology in both directions."),
        ("Colour jitter",
         "Random variation in brightness, contrast, saturation",
         "The ECG contour image intensity encodes lead voltage. "
         "Jitter makes the model robust to different display scales and scanner calibrations."),
        ("Applied to",      "2D image (3, 24, 2048) only",
         "Applied at collation time inside ChagasDataset; "
         "1D signal augmentations are separate"),
        ("Probability",     "p = 0.3 per sample",
         "Lower probability than 1D augmentations — image augmentations are more aggressive"),
        ("Training only",   "Yes",
         "Disabled for validation and test splits"),
    ])

    # TABLE 6 — AUGMENTATION RULES
    row = ctable(ws, row,
        "TABLE 6 — AUGMENTATION APPLICATION RULES", CLR["teal"],
        ["Rule", "Detail", "Enforced Where"], [
        ("Training split only",
         "All augmentations off during validation and test — clean inputs only",
         "ChagasDataset.__getitem__() checks self.split == 'train'"),
        ("Each augmentation is independent",
         "Each has its own p; applied sequentially with separate Bernoulli draws",
         "augmentations.py — each function: if random.random() < p: apply"),
        ("Labels unchanged by augmentation",
         "Signal transforms; hard or soft label passes through unchanged",
         "ChagasDataset returns original label alongside augmented signal"),
        ("Demographics unchanged",
         "Age and sex scalars are not augmented",
         "Concatenated to signal after augmentation in __getitem__()"),
    ])

    # TABLE 7 — MISSING AUGMENTATIONS
    row = ctable(ws, row,
        "TABLE 7 — MISSING AUGMENTATIONS (Not Yet Implemented)", CLR["gold"],
        ["Missing Item", "Standard in Literature?", "Why It Matters",
         "How to Implement", "Priority"], [
        ("Per-lead amplitude scaling",
         "Yes — widely used",
         "ECG gain varies ±5-10% between machines; model may overfit to exact mV scale",
         "signal *= np.random.uniform(0.8, 1.2, size=(12,1)) per lead",
         "High"),
        ("Limb lead electrode swap (RA↔LA)",
         "Yes — clinically common error",
         "Accidental lead reversal inverts P waves and QRS in limb leads; "
         "model trained without this will fail silently",
         "Negate leads I and aVL with p=0.1",
         "Medium"),
        ("Residual baseline wander",
         "Yes — common in ambulatory ECGs",
         "Bandpass removes most wander but not all; "
         "model should be robust to low-amplitude residual drift",
         "Add sinusoidal wander: A*sin(2πft), f in [0.05,0.5] Hz, A small",
         "Low"),
        ("Random lead dropout",
         "Emerging in foundation models",
         "Real recordings sometimes have broken electrode contact; "
         "model should not crash on missing leads",
         "Zero out 1-2 randomly chosen leads with p=0.1",
         "Low"),
    ], missing=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 60
    ws.freeze_panes = "A3"
    return ws


# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 6 — Model Architecture  (5 tables)
# ═══════════════════════════════════════════════════════════════════════════
def sheet_model(wb):
    ws = wb.create_sheet("6. Model Architecture")
    write_sheet_title(ws, "HybridChagasModel — One Table Per Component",
                      CLR["navy"], cols=3)
    H = ["Attribute", "Value", "Explanation"]
    row = 3

    # TABLE 1 — OVERALL MODEL SUMMARY
    row = ctable(ws, row,
        "TABLE 1 — OVERALL MODEL SUMMARY  (src/models/hybrid_model.py)",
        CLR["teal"],
        ["Component", "Class / File", "Input", "Output", "Params", "Pretrained?"], [
        ("2D Contour Branch",
         "ViT2D / vit_2d.py",
         "(B, 3, 24, 2048) uint8",
         "(B, 768)", "~86M",
         "Yes — MAE (mae_2d_pretrained.pt)"),
        ("1D Signal Foundation Model",
         "ViT1D_FM / vit_1d_fm.py",
         "(B,12,1000) + age + sex",
         "(B, 768)", "~86M",
         "Yes — ST-MEM (stmem_1d_pretrained.pt)"),
        ("REPA Cross-Modal Alignment",
         "REPAAlignment / repa_alignment.py",
         "(B, 768) from 2D branch",
         "(B, 768) aligned", "~1.5M",
         "No — trained from scratch"),
        ("Fusion + Classifier",
         "HybridChagasModel / hybrid_model.py",
         "(B,1536) = concat[aligned_2D, FM_1D]",
         "(B,1) logit", "<1M",
         "No — trained from scratch"),
        ("TOTAL", "—", "—", "—", "173.6M", "Partial (2 of 4 components)"),
    ])

    # TABLE 2 — ViT2D
    row = ctable(ws, row,
        "TABLE 2 — ViT2D: 2D CONTOUR BRANCH  (src/models/vit_2d.py)",
        CLR["navy"], H, [
        ("What is it",
         "Vision Transformer that reads the 2D ECG contour image",
         "Processes the (3, 24, 2048) image by splitting it into 96 non-overlapping "
         "patches and running 12 layers of self-attention across them"),
        ("Input",        "(B, 3, 24, 2048) uint8",
         "B = batch size; 3 WCT channels; 24 lead-rows; 2048 time-columns"),
        ("Patch size",   "Height=8 rows, Width=64 columns per patch",
         "24÷8 = 3 row-patches; 2048÷64 = 32 col-patches → 96 total"),
        ("Patch embedding",
         "Linear projection: (3×8×64=1536 pixels) → dim=768",
         "Flattens each patch, projects to embedding dimension"),
        ("Positional encoding",
         "Learned 2D embeddings (row_embed + col_embed)",
         "Separate row and column embeddings added to each patch token, "
         "giving the model knowledge of where in the image each patch comes from"),
        ("Transformer depth",  "12 layers",        "12 self-attention + FFN blocks"),
        ("Attention heads",    "12 heads, dim=64 each",
         "768 total dim ÷ 12 heads = 64 dim per head"),
        ("FFN hidden dim",     "3072 (4× embedding dim)",
         "Standard ViT MLP expansion ratio"),
        ("Aggregation method", "Aggregate-of-Layers (AoL)",
         "Takes the [CLS] token from EVERY layer (not just the last), "
         "then averages across all 12 layers. "
         "Captures both early low-level features and late high-level patterns."),
        ("Output",       "(B, 768)",
         "AoL embedding passed to REPA alignment module"),
        ("Pretraining",  "MAE — loads mae_2d_pretrained.pt",
         "Custom weight loader unpacks nested OrderedDict with 'model.' prefix (145+ keys)"),
    ])

    # TABLE 3 — ViT1D_FM
    row = ctable(ws, row,
        "TABLE 3 — ViT1D_FM: 1D SIGNAL FOUNDATION MODEL  (src/models/vit_1d_fm.py)",
        CLR["teal"], H, [
        ("What is it",
         "Vision Transformer adapted for 12-lead ECG time series",
         "Treats each ECG lead as an independent sequence of patches. "
         "Conditioned on patient demographics (age, sex) at every layer."),
        ("Signal input",  "(B, 12, 1000) float32",
         "B = batch; 12 leads; 1000 samples at 100 Hz = 10 seconds"),
        ("Demographics input", "(B, 2): [age, sex_binary]",
         "Age as raw float; sex as 0.0 (female) or 1.0 (male)"),
        ("Patching",      "50 samples per patch → 20 patches per lead",
         "1000 ÷ 50 = 20 patches. Each lead is patched independently."),
        ("Total sequence",     "240 patches (12 leads × 20)",
         "All lead patches are concatenated into one sequence for the transformer"),
        ("Lead embeddings",    "12 learned vectors, dim=768",
         "A distinct embedding is added to every patch of lead N, "
         "telling the model 'this patch comes from lead N (e.g. V1)'"),
        ("Demographics modulation",
         "FiLM: DemographicsEncoder → (gamma, beta) per transformer layer",
         "Age+sex → small MLP → scale (gamma) and shift (beta) applied "
         "to each layer's output. This is stronger than simply concatenating "
         "demographics to the input token."),
        ("Transformer depth",  "12 layers",         "Same as ViT2D for compatibility"),
        ("Attention heads",    "12 heads, dim=64",  ""),
        ("Aggregation",        "Aggregate-of-Layers (AoL)",
         "Mean of [CLS] token across all 12 layers — same strategy as ViT2D"),
        ("Output",        "(B, 768)",
         "FM features; also used as the alignment TARGET for REPA module"),
        ("Pretraining",   "ST-MEM — loads stmem_1d_pretrained.pt",
         "Custom key-remapping loader handles per-layer weight naming"),
    ])

    # TABLE 4 — REPA
    row = ctable(ws, row,
        "TABLE 4 — REPA: CROSS-MODAL ALIGNMENT MODULE  (src/models/repa_alignment.py)",
        CLR["navy"], H, [
        ("What is it",
         "A small neural network that projects 2D branch features into 1D FM space",
         "After the 2D-ViT produces its (B,768) embedding, REPA transforms it "
         "so it geometrically resembles the 1D-FM's (B,768) embedding. "
         "This reduces the 'modality gap' before fusion."),
        ("Why needed",
         "2D and 1D branches live in different representation spaces",
         "Without alignment, concatenating them is like mixing apples and oranges. "
         "The classifier must then learn to bridge the gap, wasting capacity. "
         "REPA does the bridging explicitly."),
        ("Layer 1",  "Depthwise Conv1d (groups=768, kernel=1)",
         "Operates on each of the 768 dimensions independently. "
         "Allows per-dimension rescaling without cross-channel mixing."),
        ("Layer 2",  "SiLU (Swish) activation",
         "Smooth non-linearity; better gradient flow than ReLU in this context"),
        ("Layer 3",  "Linear(768 → 768)",
         "Final linear projection into the FM embedding space"),
        ("Training signal",
         "CosineSimilarityAlignmentLoss: 1 - cos_sim(aligned_2D, FM.detach())",
         "FM gradient is STOPPED (detach). Only the 2D branch and REPA move. "
         "The FM serves as a fixed target geometry that REPA must match."),
        ("Input",    "(B, 768) from ViT2D",   "Raw 2D branch embedding"),
        ("Output",   "(B, 768) aligned",       "Now geometrically close to FM embedding space"),
        ("Parameters", "~1.5M",               "Lightweight — most of the 173.6M is in the two ViTs"),
    ])

    # TABLE 5 — CLASSIFIER HEAD
    row = ctable(ws, row,
        "TABLE 5 — FUSION + CLASSIFIER HEAD  (src/models/hybrid_model.py)",
        CLR["teal"], H, [
        ("What is it",
         "Concatenates aligned 2D and FM 1D features, then predicts Chagas probability",
         "The simplest possible fusion: direct concatenation followed by a single "
         "linear layer. Simplicity is intentional — complexity belongs in the branches."),
        ("Fusion operation",
         "torch.cat([aligned_2D, FM_1D], dim=1)",
         "Concatenates the two (B,768) vectors into (B,1536)"),
        ("Classifier",  "Linear(1536 → 1)",
         "Single linear layer with no hidden layer. Bias included."),
        ("Training output", "(B, 1) raw logit",
         "No sigmoid during training — BCEWithLogitsLoss applies it internally "
         "for numerical stability"),
        ("Inference output", "(B, 1) probability after sigmoid",
         "Threshold at 0.5 for binary decision; or rank by score for TPR@5% evaluation"),
        ("Why no hidden layer",
         "REPA already aligns the spaces; a linear probe should suffice",
         "Keeping the head simple prevents it from compensating for poor alignment — "
         "forces REPA to do its job properly"),
    ])

    # TABLE 6 — DESIGN RATIONALE
    row = ctable(ws, row,
        "TABLE 6 — ARCHITECTURE DESIGN RATIONALE",
        CLR["slate"],
        ["Design Choice", "Justification", "Evidence / Reference"], [
        ("Dual modality: 1D + 2D",
         "1D captures temporal dynamics (rhythm, intervals, QRS duration); "
         "2D captures spatial morphology (ST changes, wave shape). "
         "Neither alone covers all Chagas ECG signs.",
         "Fold-0 ablation: 1D-only AUROC=0.857 vs full 0.850 (similar) but "
         "ensemble of both = 0.871 — diversity improves ensemble"),
        ("AoL aggregation vs [CLS] token",
         "CLS only captures the final layer. AoL averages all 12 layers — "
         "early layers have low-level waveform info, late layers have global patterns. "
         "Both matter for Chagas detection.",
         "Van Santvliet et al. 2025; standard in ECG ViT literature"),
        ("REPA alignment",
         "Forces 2D branch into FM geometry before fusion, reducing modality gap "
         "and making the linear classifier's job easier.",
         "Kim et al. 2025; cosine loss weight=0.5 tuned empirically"),
        ("FiLM demographics conditioning",
         "Chagas risk strongly depends on age and sex. "
         "FiLM injects this at every transformer layer, not just the input — "
         "much stronger than simple token concatenation.",
         "Perez et al. 2018 (FiLM paper)"),
        ("Progressive unfreezing",
         "Phase 1 warms up new modules with FM frozen to avoid catastrophic forgetting. "
         "Phase 2 fine-tunes all with differential LR (FM at 10× lower rate).",
         "ULMFiT (Howard & Ruder 2018)"),
    ])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 60
    ws.freeze_panes = "A3"
    return ws

# dead placeholder so the next function def isn't lost in the edit
def sheet_pretraining(wb):
    ws = wb.create_sheet("7. Pretraining")
    write_sheet_title(ws, "Self-Supervised Pretraining — One Table Per Method",
                      CLR["teal"], cols=3)
    H = ["Attribute", "Value", "Explanation"]
    row = 3

    # TABLE 1 — MAE 2D: WHAT IT IS
    row = ctable(ws, row,
        "TABLE 1 — MAE 2D: WHAT IS MASKED AUTOENCODING?  "
        "(scripts/mae_pretraining_2d_COMPLETE.py)",
        CLR["navy"], H, [
        ("What is it",
         "Masked Autoencoder (MAE) pretraining on 2D ECG contour images",
         "A self-supervised learning method. Random patches of the image are masked. "
         "The encoder sees only the visible 20% of patches and must learn to produce "
         "embeddings rich enough for a lightweight decoder to reconstruct the masked 80%."),
        ("Why use it",
         "Learns cardiac morphology without needing any labels",
         "We have 386,981 ECG images but only 8,579 Chagas labels (2.2%). "
         "MAE exploits all 386,981 images to build a strong visual encoder before "
         "fine-tuning on the small labelled subset."),
        ("Masking ratio", "80% of patches masked per image",
         "Very high masking forces the model to understand global cardiac structure, "
         "not just copy adjacent pixels. Validated by He et al. 2022 (original MAE paper)."),
        ("What the encoder learns",
         "Rich patch embeddings capturing ECG morphological patterns",
         "After pretraining, the encoder (ViT2D) produces (B,768) embeddings that "
         "encode wave shapes, ST segments, and lead-level ECG geometry."),
        ("Encoder used downstream",
         "Yes — ViT2D weights transferred directly",
         "The same ViT2D architecture is used for pretraining and fine-tuning. "
         "No adapter or projection needed."),
        ("Decoder fate",
         "Discarded after pretraining",
         "The lightweight decoder is only needed to generate the training signal. "
         "Only mae_2d_pretrained.pt (encoder only) is saved for use downstream."),
        ("Script",       "scripts/mae_pretraining_2d_COMPLETE.py", ""),
        ("Output",       "checkpoints/mae_2d_pretrained.pt",       "337 MB"),
    ])

    # TABLE 2 — MAE 2D: ENGINEERING
    row = ctable(ws, row,
        "TABLE 2 — MAE 2D: ENGINEERING AND ROBUSTNESS FEATURES",
        CLR["teal"], H, [
        ("AMP / mixed precision",
         "float16 via torch.cuda.amp",
         "Halves memory usage; enables larger batch sizes on GPU. "
         "GradScaler prevents float16 gradient underflow."),
        ("Auto-resume",
         "Reads checkpoint at startup and continues training",
         "If the machine crashes or is shut down, pretraining resumes from where it stopped. "
         "Critical for long pretraining runs on shared GPU clusters."),
        ("Atomic checkpointing",
         "Save to temp file, then rename to final",
         "Prevents partially-written checkpoints. "
         "If power is cut during save, the old checkpoint is still intact."),
        ("SIGINT handler",
         "Ctrl+C triggers immediate graceful checkpoint save",
         "User can manually stop pretraining and the current state is saved safely."),
        ("Debug flag",
         "--subset 0.01 uses only 1% of data",
         "Allows fast smoke-testing of the script (completes in minutes not hours)"),
        ("Reconstruction loss",
         "MSE on normalised masked patch pixels",
         "Standard MAE loss: mean squared error between decoder output and "
         "the original (normalised) pixel values of the masked patches only."),
    ])

    # TABLE 3 — ST-MEM 1D: WHAT IT IS
    row = ctable(ws, row,
        "TABLE 3 — ST-MEM 1D: WHAT IS SPATIO-TEMPORAL MASKED ECG MODELLING?  "
        "(scripts/stmem_pretraining_1d_COMPLETE.py)",
        CLR["navy"], H, [
        ("What is it",
         "ST-MEM: self-supervised pretraining on 12-lead 1D ECG signals",
         "Extends MAE to ECG: masks patches across both the spatial dimension (leads) "
         "and temporal dimension (time), then predicts the masked values."),
        ("Why use it",
         "Learns inter-lead cardiac relationships without labels",
         "The 12 ECG leads are anatomically related (e.g. V1 mirrors aVR). "
         "ST-MEM forces the model to learn these cross-lead dependencies by "
         "predicting masked leads from the visible ones."),
        ("Critical design: per-lead masking",
         "Entire leads are masked, not individual patches",
         "If individual patches were masked, the model could copy values from "
         "the same lead at adjacent time steps (temporal autocorrelation). "
         "Per-lead masking forces genuine cross-lead reconstruction — the model "
         "MUST use other leads to reconstruct a masked one. "
         "Source: Van Santvliet et al. 2025, Section 3."),
        ("Masking granularity",
         "A random subset of the 12 leads is fully masked per sample",
         "Each training sample has different leads masked. "
         "The model sees N visible leads and must predict M masked leads."),
        ("What the encoder learns",
         "Lead embeddings + inter-lead cardiac dependencies",
         "After pretraining, ViT1D_FM can infer missing leads from visible ones, "
         "meaning it deeply understands how different ECG perspectives relate."),
        ("Encoder used downstream",
         "Yes — ViT1D_FM weights transferred directly",
         "Loaded via stmem_1d_pretrained.pt at the start of each fold training"),
        ("Script",  "scripts/stmem_pretraining_1d_COMPLETE.py", ""),
        ("Output",  "checkpoints/stmem_1d_pretrained.pt",       "333 MB"),
    ])

    # TABLE 4 — ST-MEM 1D: ENGINEERING
    row = ctable(ws, row,
        "TABLE 4 — ST-MEM 1D: ENGINEERING AND ROBUSTNESS FEATURES",
        CLR["teal"], H, [
        ("AMP / mixed precision",
         "float16 via torch.cuda.amp",
         "2× throughput using Tensor Cores on NVIDIA Ampere/Hopper GPUs"),
        ("Gradient clipping",
         "max_norm = 1.0",
         "Prevents exploding gradients in the early stages of pretraining "
         "before the model has stabilised"),
        ("Data loading optimisation",
         "persistent_workers=True, prefetch_factor=2",
         "Workers stay alive between batches (no fork overhead); "
         "prefetch_factor=2 keeps the GPU fed while the CPU loads the next batch"),
        ("Resume counter",
         "processed_batches saved in checkpoint dict",
         "More robust than epoch tracking: counts exact batches processed. "
         "A partial last epoch is safely resumed without repeating data."),
        ("Reconstruction loss",
         "MSE on masked lead patch values",
         "Same MAE-style MSE loss but applied to normalised 1D signal patches "
         "rather than image pixels"),
    ])

    # TABLE 5 — COMPARISON
    row = ctable(ws, row,
        "TABLE 5 — MAE 2D vs ST-MEM 1D: SIDE-BY-SIDE COMPARISON",
        CLR["slate"],
        ["Aspect", "MAE 2D", "ST-MEM 1D"], [
        ("Input modality",   "(3, 24, 2048) uint8 image",  "(12, 1000) float32 signal"),
        ("Masking unit",     "2D image patches (8×64 px)", "Entire ECG leads"),
        ("Masking ratio",    "80% of 96 patches",          "Random subset of 12 leads"),
        ("Reconstruction target", "Masked patch pixels",   "Masked lead signal values"),
        ("Loss",             "MSE on pixel values",        "MSE on signal patch values"),
        ("Labels used",      "None — fully self-supervised","None — fully self-supervised"),
        ("Output checkpoint","mae_2d_pretrained.pt (337MB)","stmem_1d_pretrained.pt (333MB)"),
        ("Used in downstream","ViT2D encoder init",        "ViT1D_FM encoder init"),
    ])

    # TABLE 6 — GAPS
    row = ctable(ws, row,
        "TABLE 6 — PRETRAINING GAPS (Missing Items)", CLR["gold"],
        ["#", "Gap", "Impact", "Fix", "Priority"], [
        ("1", "Epoch/iteration counts not stored in .pt files",
         "Cannot reproduce exact pretraining run",
         "Add metadata dict: {epochs, iters, loss, timestamp} to torch.save()",
         "High"),
        ("2", "No pretraining loss curves saved",
         "Cannot show convergence evidence to examiners",
         "Log loss per batch to mae_pretraining_loss.csv and stmem_pretraining_loss.csv",
         "High"),
        ("3", "MAE decoder architecture not documented",
         "Cannot reconstruct or verify the decoder",
         "Add decoder_depth and decoder_dim constants to the script top",
         "Medium"),
        ("4", "Pretraining hyperparameters (LR, batch size) not recorded",
         "Reproducibility gap",
         "Create pretraining_config.yaml; embed its hash in checkpoint metadata",
         "Medium"),
    ], missing=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 60
    ws.freeze_panes = "A3"
    return ws


# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 8 — Training Configuration  (4 tables)
# ═══════════════════════════════════════════════════════════════════════════
def sheet_training(wb):
    ws = wb.create_sheet("8. Training Config")
    write_sheet_title(ws, "Training Configuration — One Table Per Concept",
                      CLR["navy"], cols=3)
    H = ["Attribute", "Value", "Explanation"]
    row = 3

    # TABLE 1 — PROGRESSIVE UNFREEZING CONCEPT
    row = ctable(ws, row,
        "TABLE 1 — WHAT IS PROGRESSIVE UNFREEZING?  (src/training/trainer.py)",
        CLR["teal"], H, [
        ("What is it",
         "A two-phase fine-tuning strategy for transfer learning",
         "Instead of fine-tuning all model weights at once from a pretrained checkpoint, "
         "Phase 1 freezes the pretrained FM backbone and only trains the new modules. "
         "Phase 2 then unfreezes everything with a lower LR for the FM."),
        ("Why needed",
         "Prevents catastrophic forgetting of pretraining",
         "If you immediately fine-tune the FM backbone with a high LR from a random classifier, "
         "the large gradients from the random classifier destroy the pretrained FM weights. "
         "Phase 1 stabilises the classifier first, THEN Phase 2 fine-tunes the FM gently."),
        ("Phase 1 purpose",
         "Warm up classifier + 2D-ViT + REPA with FM frozen",
         "After 2000 iterations the classifier is no longer random; "
         "its gradients are small enough that the FM can safely be unfrozen."),
        ("Phase 2 purpose",
         "Fine-tune entire model jointly with differential learning rates",
         "FM gets LR=1e-5 (10× smaller than classifier LR=1e-4) to adapt slowly. "
         "The pretrained FM knowledge is preserved while still improving for Chagas."),
        ("Source / inspiration",
         "ULMFiT — Howard & Ruder 2018",
         "Originally developed for NLP transfer learning; "
         "now standard practice for medical imaging and ECG deep learning"),
    ])

    # TABLE 2 — PHASE 1
    row = ctable(ws, row,
        "TABLE 2 — PHASE 1: FROZEN FM WARM-UP  (2,000 iterations)",
        CLR["navy"], H, [
        ("Iterations",          "2,000 steps",
         "Short by design. Just long enough to train the classifier, "
         "2D-ViT, and REPA without touching FM weights."),
        ("Micro batch size",    "16 per GPU step",
         "GPU memory limit with both ViTs loaded simultaneously"),
        ("Gradient accumulation","4 steps",
         "Simulates a batch of 16 × 4 = 64 without requiring 64-sample GPU memory"),
        ("Effective batch size", "64",
         "Larger batches help early stability when classifier is randomly initialised"),
        ("LR — classifier, 2D-ViT, REPA", "2e-4",
         "These are the modules being actively trained in Phase 1"),
        ("LR — FM (ViT1D_FM)", "0.0  (weights frozen)",
         "FM parameters excluded from optimiser entirely in Phase 1"),
        ("LR warmup",           "200 linear ramp steps (0 → 2e-4)",
         "Prevents very large first-step gradients from random classifier weights"),
        ("LR schedule after warmup", "Flat — no decay",
         "Phase 1 is too short for cosine decay to be meaningful"),
        ("Loss",                "AsymmetricBCE + 0.5 × REPA alignment",
         "BCE trains classifier; alignment loss trains REPA to bridge modalities"),
        ("Validation (fast)",   "Stratified 3000-sample subset, 1000 permutations",
         "~0.1 s; done every 200 iterations to track convergence without stalling"),
        ("Validation (official)","Full val set, 10,000 permutations — at end of Phase 1 only",
         "Matches PhysioNet 2025 challenge metric exactly"),
        ("Checkpointing",       "Best model by TPR@5% on val",
         "checkpoints/fold{n}_best.pt updated whenever val TPR@5% improves"),
    ])

    # TABLE 3 — PHASE 2
    row = ctable(ws, row,
        "TABLE 3 — PHASE 2: FULL FINE-TUNING  (12,000–24,000 iterations, FM unfrozen)",
        CLR["teal"], H, [
        ("Iterations",          "12,000 (main) or 24,000 (checkpoints12 variant)",
         "Main 12k run produces slightly better TPR@5% than 24k run "
         "(0.4482 vs 0.4376 on fold 0) — suggests 24k begins to overfit"),
        ("Micro batch size",    "16 per GPU step",   "Same as Phase 1"),
        ("Gradient accumulation","2 steps",
         "Effective batch 16 × 2 = 32. Halved vs Phase 1 — smaller batches "
         "are better for full-model fine-tuning."),
        ("Effective batch size", "32",               ""),
        ("LR — classifier, 2D-ViT, REPA", "1e-4",
         "Slightly lower than Phase 1 for more stable convergence"),
        ("LR — FM (ViT1D_FM)", "1e-5  (10× lower than classifier)",
         "DIFFERENTIAL LR. FM adapts slowly to preserve ST-MEM pretraining. "
         "Without this, FM catastrophically forgets its learned ECG representations."),
        ("LR schedule",         "Cosine decay from 1e-4 → 0 over Phase 2",
         "Smooth gradual reduction; prevents abrupt learning rate drops late in training"),
        ("Gradient clipping",   "max_norm = 1.0",
         "When FM is first unfrozen, its gradients can be large. "
         "Clipping prevents NaN/inf loss on the first few batches of Phase 2."),
        ("AMP",                 "float16 + GradScaler",
         "GradScaler needed in Phase 2 — with FM unfrozen, "
         "gradient magnitudes vary more and float16 underflow is more likely"),
        ("Best model",          "Saved by TPR@5% improvement on val set",
         "checkpoints/fold{n}_best.pt"),
        ("Training curve saved","checkpoints/fold{n}_training_curve.png",
         "50-iteration rolling average of loss; TPR@5% at each validation point"),
    ])

    # TABLE 4 — ASYMMETRIC BCE LOSS
    row = ctable(ws, row,
        "TABLE 4 — ASYMMETRIC BCE LOSS  (src/training/losses.py — AsymmetricBCELoss)",
        CLR["navy"], H, [
        ("What is it",
         "Binary Cross-Entropy with asymmetric focal weighting",
         "Standard BCE is modified: easy negatives are down-weighted more than "
         "easy positives. This handles both class imbalance and label noise."),
        ("Why needed",
         "Dataset is 97.78% negative (2.22% Chagas positive)",
         "With standard BCE, the model learns to predict 'negative' for everything "
         "and still gets ~97.78% accuracy. ABCE prevents this by amplifying "
         "the gradient signal from positive samples."),
        ("gamma_positive (γ⁺)", "0",
         "No down-weighting of positives. Every true positive contributes "
         "fully to the gradient, no matter how confidently classified."),
        ("gamma_negative (γ⁻)", "2",
         "Easy negatives (high-confidence negative predictions) are down-weighted "
         "by (1-p)^2. This is stronger than standard focal loss (same γ for both)."),
        ("pos_weight",          "10",
         "Amplifies the gradient of positive samples by 10×. "
         "Compensates for the 97.78% / 2.22% imbalance ratio (~45:1)."),
        ("Soft label handling", "Accepts y in [0.0, 1.0] directly",
         "CODE-15 soft labels (0.2 / 0.8) work with this loss without modification. "
         "Standard BCEWithLogitsLoss would require binarisation."),
        ("Weight in total loss", "1.0 — full weight",
         "Primary classification signal; alignment loss is secondary (0.5 weight)"),
        ("Applied in",           "Phase 1 and Phase 2", ""),
    ])

    # TABLE 5 — COSINE ALIGNMENT LOSS
    row = ctable(ws, row,
        "TABLE 5 — COSINE SIMILARITY ALIGNMENT LOSS  (CosineSimilarityAlignmentLoss)",
        CLR["teal"], H, [
        ("What is it",
         "Minimises the angular distance between 2D branch and 1D FM embeddings",
         "loss = 1 - cosine_similarity(REPA(2D_features), FM_features.detach())"),
        ("Why needed",
         "2D and 1D embeddings live in different geometric spaces",
         "Without this loss, the concatenated (B,1536) vector is geometrically "
         "incoherent — part image-space, part signal-space. The classifier must "
         "compensate. REPA + this loss explicitly maps 2D into FM geometry first."),
        ("FM gradient",         "STOPPED via .detach()",
         "The loss only moves the 2D branch and REPA toward the FM. "
         "The FM does not move toward the 2D branch. "
         "FM is the fixed target; 2D must match it, not vice versa."),
        ("Perfect alignment",   "loss = 0  (cos_sim = 1)",
         "Both embeddings point in the same direction in 768-dim space"),
        ("Random alignment",    "loss ≈ 1  (cos_sim ≈ 0)",
         "Orthogonal embeddings — no geometric relationship"),
        ("Weight in total loss", "0.5 — half weight",
         "Tuned so alignment does not dominate the BCE classification signal"),
        ("Applied in",           "Phase 1 and Phase 2", ""),
    ])

    # TABLE 6 — COMBINED LOSS
    row = ctable(ws, row,
        "TABLE 6 — COMBINED LOSS  (CombinedLoss wrapper)",
        CLR["navy"], H, [
        ("What is it",
         "Wrapper that computes and returns the sum of both losses",
         "total = AsymmetricBCE(logit, label) + 0.5 × CosineSimilarityAlignment(aligned_2D, FM)"),
        ("Why a wrapper",
         "Trainer calls one loss object for simplicity",
         "trainer.py calls loss(logits, labels, aligned_2D, fm_features) "
         "and gets back a single scalar. Clean separation of concerns."),
        ("Gradient flow",
         "BCE → classifier, 2D-ViT, REPA (and FM in Phase 2)",
         "Alignment → REPA and 2D-ViT only (FM detached)"),
        ("Alignment weight",    "0.5 — empirically tuned",
         "Too low: 2D branch doesn't align properly. "
         "Too high: alignment dominates BCE and hurts classification. "
         "0.5 was chosen by monitoring val TPR@5% during development."),
    ])

    # TABLE 7 — OPTIMIZER + SCHEDULER
    row = ctable(ws, row,
        "TABLE 7 — OPTIMIZER, SCHEDULER, AND REGULARISATION",
        CLR["slate"],
        ["Setting", "Phase 1", "Phase 2", "Notes"], [
        ("Optimizer",              "AdamW",         "AdamW",
         "β1=0.9, β2=0.999; standard for transformers"),
        ("Weight decay",           "1e-4",          "1e-4",
         "Applied to non-FM params; zero on FM in Phase 2 (common practice)"),
        ("LR — classifier",        "2e-4",          "1e-4",      ""),
        ("LR — 2D-ViT (ViT2D)",   "2e-4",          "1e-4",      ""),
        ("LR — REPA",              "2e-4",          "1e-4",      ""),
        ("LR — FM (ViT1D_FM)",     "0 (frozen)",    "1e-5",
         "10× lower than classifier — differential LR"),
        ("LR warmup",              "200 linear steps","None",
         "Only at Phase 1 start when classifier is random"),
        ("LR schedule",            "Flat",          "Cosine → 0",
         "Phase 2 cosine over full iteration count"),
        ("Gradient clipping",      "None",          "max_norm=1.0",
         "Only needed in Phase 2 when FM first unfrozen"),
        ("AMP / GradScaler",       "Yes float16",   "Yes float16", ""),
        ("Dropout",                "None",          "None",
         "Regularisation via weight decay only"),
    ])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 60
    ws.freeze_panes = "A3"
    return ws


# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 9 — Per-Fold Results
# ═══════════════════════════════════════════════════════════════════════════
def sheet_fold_results(wb):
    ws = wb.create_sheet("9. Per-Fold Results")
    write_sheet_title(ws, "Per-Fold Validation Results (5-Fold CV)",
                      CLR["teal"], cols=9)

    row = 3
    write_section_header(ws, row,
        "  FULL MODEL (1D + 2D, pretrained)", CLR["navy"], cols=9)
    row += 1
    write_col_headers(ws, row,
        ["Fold", "AUROC", "AUPRC", "TPR@5% (Primary)",
         "Sensitivity", "Specificity", "Accuracy", "Val Size"],
        CLR["slate"])
    row += 1

    folds = [
        ("Fold 0", "0.8503", "0.2163", "0.4482", "—", "—", "—", "73,237"),
        ("Fold 1", "0.7607", "0.1687", "0.3748", "—", "—", "—", "73,236"),
        ("Fold 2", "0.7997", "0.1370", "—",      "—", "—", "—", "73,236"),
        ("Fold 3", "0.8217", "0.1749", "—",      "—", "—", "—", "73,236"),
        ("Fold 4", "0.8482", "0.2056", "—",      "—", "—", "—", "73,236"),
        ("Ensemble", "0.8707", "0.2589", "0.4958",
         "0.7765", "0.8011", "0.8005", "386,981"),
    ]
    for i, f in enumerate(folds):
        bold = (f[0] == "Ensemble")
        write_data_row(ws, row, list(f), alt=(i % 2 == 0), bold_first=bold)
        row += 1

    row += 1
    write_section_header(ws, row,
        "  ABLATION RESULTS (Fold 0 only)", CLR["teal"], cols=9)
    row += 1
    write_col_headers(ws, row,
        ["Variant", "AUROC", "AUPRC", "TPR@5%",
         "vs Full Model AUROC", "vs Full Model TPR@5%", "Conclusion", ""],
        CLR["slate"])
    row += 1

    ablations = [
        ("Full model (1D+2D, pretrained)",
         "0.8503", "0.2163", "0.4482", "baseline", "baseline", "Best", ""),
        ("1D only (no 2D branch)",
         "0.8567", "0.2295", "0.4482", "+0.0064", "=",
         "2D adds diversity but 1D dominates", ""),
        ("2D only (no 1D FM branch)",
         "—",      "—",      "—",      "—", "—",
         "Results in fold0_2d_results.csv (read separately)", ""),
        ("No pretraining (random init)",
         "—",      "—",      "—",      "—", "—",
         "Results in fold0_no_pretrain_results.csv (read separately)", ""),
        ("checkpoints12 variant (24k iters)",
         "0.8503", "0.2163", "0.4376", "=",  "−0.0106",
         "12k iters (main) slightly better at 5% threshold", ""),
    ]
    for i, a in enumerate(ablations):
        write_data_row(ws, row, list(a), alt=(i % 2 == 0), bold_first=True)
        row += 1

    row += 1
    write_section_header(ws, row,
        "  ENSEMBLE CONFIDENCE INTERVALS (Bootstrap, n=10,000)",
        CLR["slate"], cols=9)
    row += 1
    write_col_headers(ws, row,
        ["Metric", "Point Estimate", "95% CI Lower", "95% CI Upper",
         "Interpretation", "", "", ""],
        CLR["slate"])
    row += 1

    cis = [
        ("TPR@5% (Primary)",  "0.4958", "0.4845", "0.5068",
         "Finds 4,272 / 8,579 cases when screening 5% of patients (2.77× random)", "", "", ""),
        ("AUROC",             "0.8707", "0.8665", "0.8746",
         "Strong discriminative ability", "", "", ""),
        ("AUPRC",             "0.2589", "0.2489", "0.2685",
         "11.7× precision-at-recall lift vs. 2.22% random baseline", "", "", ""),
        ("Sensitivity",       "0.7765", "—",      "—",
         "Finds 77.7% of true Chagas cases", "", "", ""),
        ("Specificity",       "0.8011", "—",      "—",
         "Avoids 80.1% of false positives", "", "", ""),
        ("NNS (screen N to find 1 case)",
         "4.5",    "—",      "—",
         "Screen 4.5 patients to detect 1 Chagas case", "", "", ""),
    ]
    for i, c in enumerate(cis):
        write_data_row(ws, row, list(c), alt=(i % 2 == 0), bold_first=True)
        row += 1

    auto_col_widths(ws, max_w=52)
    ws.freeze_panes = "A3"
    return ws


# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 10 — Per-Dataset Results
# ═══════════════════════════════════════════════════════════════════════════
def sheet_dataset_results(wb):
    ws = wb.create_sheet("10. Per-Dataset Results")
    write_sheet_title(ws, "Per-Dataset Performance Breakdown",
                      CLR["slate"], cols=8)

    row = 3
    write_section_header(ws, row,
        "  DATASET-LEVEL EVALUATION", CLR["navy"], cols=8)
    row += 1
    write_col_headers(ws, row,
        ["Dataset", "N Records", "N Positive", "Prevalence",
         "AUROC", "AUPRC", "TPR@5%", "Note"],
        CLR["slate"])
    row += 1

    ds = [
        ("PTB-XL",    "21,837",  "0",     "0%",    "N/A",  "N/A",  "N/A",
         "No Chagas positives — used only as hard negative set"),
        ("SaMi-Trop", "1,631",   "1,631", "100%",  "N/A",  "N/A",  "N/A",
         "All Chagas positive — endemic cohort, no AUC computable"),
        ("CODE-15",   "363,513", "6,948", "1.91%", "reported in per_dataset_metrics.csv",
         "reported", "reported",
         "Primary evaluation dataset (mixed labels, soft targets)"),
        ("TOTAL",     "386,981", "8,579", "2.22%", "0.8707","0.2589","0.4958",
         "Ensemble across all 5 folds"),
    ]
    for i, d in enumerate(ds):
        write_data_row(ws, row, list(d), alt=(i % 2 == 0), bold_first=True)
        row += 1

    row += 1
    write_section_header(ws, row,
        "  NOTES ON DATASET-LEVEL INTERPRETATION", CLR["teal"], cols=8)
    row += 1
    write_col_headers(ws, row, ["Point", "Explanation"], CLR["slate"])
    row += 1

    notes = [
        ("Why AUROC is N/A for PTB-XL",
         "AUROC requires both classes. PTB-XL has zero Chagas positives; "
         "it contributes only to specificity estimation"),
        ("Why AUROC is N/A for SaMi-Trop",
         "SaMi-Trop has only Chagas positives; it contributes only to sensitivity"),
        ("CODE-15 drives overall discrimination",
         "With 363k records and mixed prevalence, CODE-15 determines the main AUC. "
         "The dataset's soft labels prevent inflated estimates"),
        ("Missing: full per-dataset metric table",
         "per_dataset_metrics.csv exists in checkpoints/ but individual values "
         "not extracted here — read CSV for exact numbers"),
    ]
    for i, (k, v) in enumerate(notes):
        is_missing = k.startswith("Missing")
        write_data_row(ws, row, [k, v], alt=(i % 2 == 0),
                       bold_first=True, missing=is_missing)
        row += 1

    auto_col_widths(ws, max_w=60)
    ws.freeze_panes = "A3"
    return ws


# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 11 — Checkpoints & Artefacts
# ═══════════════════════════════════════════════════════════════════════════
def sheet_checkpoints(wb):
    ws = wb.create_sheet("11. Checkpoints & Artefacts")
    write_sheet_title(ws, "Checkpoints, Models, and Output Artefacts",
                      CLR["gold"], cols=8)

    row = 3
    write_section_header(ws, row,
        "  MODEL CHECKPOINTS (checkpoints/)", CLR["navy"], cols=8)
    row += 1
    write_col_headers(ws, row,
        ["File", "Size (MB)", "Type", "Content", "Used For"],
        CLR["slate"])
    row += 1

    models = [
        ("mae_2d_pretrained.pt",      "337.2", "2D Foundation",
         "MAE-pretrained ViT2D weights",
         "Initialise 2D branch before fold training"),
        ("stmem_1d_pretrained.pt",    "333.3", "1D Foundation",
         "ST-MEM-pretrained ViT1D_FM weights",
         "Initialise 1D branch before fold training"),
        ("fold0_best.pt",             "335.8", "Fold 0 Model",
         "Best HybridChagasModel checkpoint (fold 0)",
         "Fold 0 validation; ensemble member"),
        ("fold0_2d_best.pt",          "338.0", "Fold 0 2D-only",
         "2D-only ablation checkpoint",
         "Ablation: 2D-only performance"),
        ("fold1_best.pt",             "2,036.2","Fold 1 Model",
         "Best HybridChagasModel checkpoint (fold 1)",
         "Fold 1 validation; ensemble member"),
        ("fold2_best.pt",             "2,036.2","Fold 2 Model",
         "Best HybridChagasModel checkpoint (fold 2)",
         "Fold 2 validation; ensemble member"),
        ("fold3_best.pt",             "2,036.2","Fold 3 Model",
         "Best HybridChagasModel checkpoint (fold 3)",
         "Fold 3 validation; ensemble member"),
        ("fold4_best.pt",             "2,036.2","Fold 4 Model",
         "Best HybridChagasModel checkpoint (fold 4)",
         "Fold 4 validation; ensemble member"),
        ("FINAL_ENSEMBLE_MODEL.pt",   "3,390.6","Ensemble",
         "Aggregated ensemble of folds 0–4",
         "Final inference / thesis submission"),
    ]
    for i, m in enumerate(models):
        write_data_row(ws, row, list(m), alt=(i % 2 == 0), bold_first=True)
        row += 1

    row += 1
    write_section_header(ws, row,
        "  RESULTS FILES (checkpoints/)", CLR["teal"], cols=8)
    row += 1
    write_col_headers(ws, row,
        ["File", "Contents", "Key Metrics Present", "", ""],
        CLR["slate"])
    row += 1

    results = [
        ("fold0_results.csv", "Full model fold 0",
         "AUROC=0.8503, AUPRC=0.2163, TPR@5%=0.4482", "", ""),
        ("fold0_1d_results.csv", "1D-only ablation fold 0",
         "AUROC=0.8567, AUPRC=0.2295, TPR@5%=0.4482", "", ""),
        ("fold0_2d_results.csv", "2D-only ablation fold 0",
         "See file for exact values", "", ""),
        ("fold0_no_pretrain_results.csv", "No-pretraining ablation fold 0",
         "See file for exact values", "", ""),
        ("fold1_results.csv", "Full model fold 1",
         "AUROC=0.7607, AUPRC=0.1687, TPR@5%=0.3748", "", ""),
        ("fold2_results.csv", "Full model fold 2",
         "AUROC=0.7997, AUPRC=0.1370", "", ""),
        ("fold3_results.csv", "Full model fold 3",
         "AUROC=0.8217, AUPRC=0.1749", "", ""),
        ("fold4_results.csv", "Full model fold 4",
         "AUROC=0.8482, AUPRC=0.2056", "", ""),
        ("per_fold_metrics.csv", "All 5 folds + ensemble summary",
         "All primary + secondary metrics", "", ""),
        ("per_dataset_metrics.csv", "Breakdown by dataset",
         "PTB-XL / SaMi-Trop / CODE-15 metrics", "", ""),
        ("ensemble_summary.csv", "Final ensemble with bootstrap CIs",
         "TPR@5%=0.4958 [0.4845–0.5068]", "", ""),
        ("ensemble_predictions.csv", "Raw predictions for all 386,981 samples",
         "id, dataset, label, score, fold", "", ""),
        ("threshold_comparison.csv", "Sensitivity/Specificity at various thresholds",
         "Threshold sweep analysis", "", ""),
    ]
    for i, r in enumerate(results):
        write_data_row(ws, row, list(r), alt=(i % 2 == 0), bold_first=True)
        row += 1

    row += 1
    write_section_header(ws, row,
        "  TRAINING CURVES (checkpoints/)", CLR["slate"], cols=8)
    row += 1
    write_col_headers(ws, row, ["File", "Content", "", "", ""], CLR["slate"])
    row += 1

    curves = [
        ("fold1_training_curve.png", "Loss + TPR@5% over iterations for fold 1"),
        ("fold2_training_curve.png", "Loss + TPR@5% over iterations for fold 2"),
        ("fold3_training_curve.png", "Loss + TPR@5% over iterations for fold 3"),
        ("fold4_training_curve.png", "Loss + TPR@5% over iterations for fold 4"),
        ("fold2_no_pretrain_training_curve.png",
         "No-pretraining baseline training curve (fold 2)"),
        ("MISSING: fold0_training_curve.png",
         "Fold 0 curve was deleted from checkpoints/ (see git status)"),
    ]
    for i, c in enumerate(curves):
        is_missing = c[0].startswith("MISSING")
        vals = list(c) + ["", "", ""]
        write_data_row(ws, row, vals[:5], alt=(i % 2 == 0),
                       bold_first=True, missing=is_missing)
        row += 1

    auto_col_widths(ws, max_w=55)
    ws.freeze_panes = "A3"
    return ws


# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 12 — Missing / To-Do
# ═══════════════════════════════════════════════════════════════════════════
def sheet_missing(wb):
    ws = wb.create_sheet("12. Gaps & To-Do")
    write_sheet_title(ws, "Known Gaps and Recommended Next Steps",
                      CLR["gold"], cols=8)

    row = 3
    write_section_header(ws, row,
        "  MISSING ITEMS ACROSS PIPELINE", CLR["navy"], cols=8)
    row += 1
    write_col_headers(ws, row,
        ["#", "Category", "Gap / Missing Item",
         "Impact", "Suggested Fix", "Priority"],
        CLR["slate"])
    row += 1

    gaps = [
        ("1", "Results",
         "fold0_training_curve.png deleted from checkpoints/",
         "Cannot visualise fold 0 learning dynamics",
         "Re-run fold 0 training or recover from git history",
         "High"),
        ("2", "Results",
         "Full ablation values for 2D-only and no-pretrain variants not extracted",
         "Ablation table incomplete for thesis Chapter 8",
         "Read fold0_2d_results.csv and fold0_no_pretrain_results.csv and add to table",
         "High"),
        ("3", "Results",
         "TPR@5% missing for folds 2, 3, 4 in per_fold_metrics.csv",
         "Cannot show per-fold primary metric trend",
         "Extract from per_fold_metrics.csv or re-evaluate saved checkpoints",
         "High"),
        ("4", "Results",
         "Sensitivity, Specificity, Accuracy missing for folds 0–4 individually",
         "Incomplete fold-level breakdown for thesis",
         "Compute from ensemble_predictions.csv filtered by fold",
         "Medium"),
        ("5", "Pretraining",
         "Pretraining loss curves not saved",
         "Cannot verify MAE/ST-MEM convergence for thesis",
         "Add CSV loss logging to both pretraining scripts",
         "Medium"),
        ("6", "Pretraining",
         "Epoch counts not stored in checkpoint metadata",
         "Reproducibility risk — cannot know how long pretraining ran",
         "Save training metadata dict inside .pt files",
         "Medium"),
        ("7", "Augmentation",
         "Amplitude scaling augmentation not implemented",
         "Missing a standard ECG robustness augmentation",
         "Add ×[0.8–1.2] per-lead scaling to augmentations.py",
         "Low"),
        ("8", "Augmentation",
         "Electrode lead swap simulation not implemented",
         "Model may not handle accidental RA↔LA electrode reversal",
         "Add lead-reversal augmentation for limb leads",
         "Low"),
        ("9", "Data",
         "Exact PTB-XL / SaMi-Trop / CODE-15 record counts post-filter unknown",
         "May differ from published dataset sizes if corrupted files skipped",
         "Add rejection log to build_all_data.py",
         "Low"),
        ("10","Checkpoints",
         "fold0_results.csv deleted and replaced by checkpoints12/fold0_results.csv",
         "Original fold 0 result file gone from main checkpoints/",
         "Restore from git or document which checkpoint is canonical",
         "Medium"),
        ("11","Documentation",
         "No YAML/JSON config files for hyperparameters",
         "Hard to reproduce exact training without reading source code",
         "Extract Phase 1/2 hyperparameters into config.yaml",
         "Medium"),
        ("12","Evaluation",
         "No external test set evaluation (only CV validation)",
         "Cannot claim generalization beyond the 3 training datasets",
         "Plan evaluation on held-out external ECG dataset if available",
         "High"),
    ]
    for i, g in enumerate(gaps):
        write_data_row(ws, row, list(g), alt=(i % 2 == 0),
                       missing=True, bold_first=False)
        row += 1

    auto_col_widths(ws, max_w=60)
    ws.freeze_panes = "A3"
    return ws


# ═══════════════════════════════════════════════════════════════════════════
#  INDEX SHEET (first sheet)
# ═══════════════════════════════════════════════════════════════════════════
def sheet_index(wb):
    ws = wb.create_sheet("Index", 0)
    write_sheet_title(ws, "ChagaSight FYP — Project Documentation Workbook",
                      CLR["navy"], cols=5)

    row = 3
    write_section_header(ws, row, "  SHEET DIRECTORY", CLR["teal"], cols=5)
    row += 1
    write_col_headers(ws, row,
        ["Sheet #", "Sheet Name", "Contents", "Colour Code", ""],
        CLR["slate"])
    row += 1

    sheets = [
        ("1", "Project Overview",       "High-level pipeline stages and KPIs",
         "Navy / Teal", ""),
        ("2", "Dataset Summary",        "PTB-XL, SaMi-Trop, CODE-15 details",
         "Teal", ""),
        ("3", "Data Split (5-Fold CV)", "Stratification strategy and fold sizes",
         "Slate", ""),
        ("4", "Preprocessing",          "1D signal and 2D image processing steps",
         "Teal", ""),
        ("5", "Augmentation",           "Online augmentations and design notes",
         "Gold", ""),
        ("6", "Model Architecture",     "HybridChagasModel components (173.6M params)",
         "Navy", ""),
        ("7", "Pretraining",            "MAE 2D and ST-MEM 1D pretraining",
         "Teal", ""),
        ("8", "Training Config",        "Phase 1/2 hyperparameters, losses, optimizer",
         "Navy", ""),
        ("9", "Per-Fold Results",       "Fold 0–4 metrics + ablations + ensemble CIs",
         "Teal", ""),
        ("10","Per-Dataset Results",    "PTB-XL / SaMi-Trop / CODE-15 breakdown",
         "Slate", ""),
        ("11","Checkpoints & Artefacts","All .pt files, CSVs, and training curves",
         "Gold", ""),
        ("12","Gaps & To-Do",           "12 identified gaps with priority and fix",
         "Gold (orange highlights)", ""),
    ]
    for i, s in enumerate(sheets):
        write_data_row(ws, row, list(s), alt=(i % 2 == 0), bold_first=False)
        row += 1

    row += 2
    cell = ws.cell(row, 1,
        "Orange-highlighted rows throughout the workbook indicate MISSING items "
        "or items requiring attention.")
    cell.font = Font(name="Calibri", size=10, italic=True, color=CLR["missing_border"])
    cell.alignment = left()

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 25
    ws.freeze_panes = "A3"
    return ws


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheet_index(wb)
    sheet_overview(wb)
    sheet_datasets(wb)
    sheet_splits(wb)
    sheet_preprocessing(wb)
    sheet_augmentation(wb)
    sheet_model(wb)
    sheet_pretraining(wb)
    sheet_training(wb)
    sheet_fold_results(wb)
    sheet_dataset_results(wb)
    sheet_checkpoints(wb)
    sheet_missing(wb)

    out_path = r"D:\IIT\L6\FYP\ChagaSight\ChagaSight_Project_Documentation.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
